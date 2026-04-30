"""
services/report_service.py
Xuất báo cáo điểm danh — Excel + PDF.
"""
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass
from typing import Optional
from loguru import logger

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import app_config, report_config


# ─────────────────────────────────────────────
@dataclass
class ReportData:
    session_id:     int
    class_code:     str
    class_name:     str
    subject_name:   str
    session_date:   str
    start_time:     str
    end_time:       str
    teacher_name:   str
    total_students: int
    present_count:  int
    absent_count:   int
    late_count:     int = 0
    records:        list = None

    @property
    def attendance_rate(self) -> float:
        if self.total_students == 0:
            return 0.0
        return self.present_count / self.total_students * 100

    @property
    def title(self) -> str:
        return f"BẢNG ĐIỂM DANH — {self.subject_name}"


# ─────────────────────────────────────────────
def load_report_data(session_id: int) -> Optional[ReportData]:
    """Load dữ liệu điểm danh 1 buổi từ DB."""
    try:
        from database.repositories import session_repo, record_repo, class_repo

        session = session_repo.get_by_id(session_id)
        if not session:
            logger.error(f"Không tìm thấy session {session_id}")
            return None

        records = record_repo.get_session_report(session_id)
        present = [r for r in records if r.get("status") == "PRESENT"]
        absent  = [r for r in records if r.get("status") == "ABSENT"]

        # Lấy teacher_name từ Classes — session không join trực tiếp
        teacher_name = ""
        cls = class_repo.get_by_id(session.class_id)
        if cls:
            teacher_name = cls.teacher_name or ""

        data = ReportData(
            session_id=session_id,
            class_code=getattr(session, "class_code", "") or "",
            class_name=getattr(session, "class_name", "") or "",
            subject_name=session.subject_name,
            session_date=(
                session.session_date.strftime("%d/%m/%Y")
                if session.session_date else ""
            ),
            start_time=(
                session.start_time.strftime("%H:%M:%S")
                if session.start_time else "—"
            ),
            end_time=(
                session.end_time.strftime("%H:%M:%S")
                if session.end_time else "—"
            ),
            teacher_name=teacher_name,
            # total_students = len(records) — không dùng session.total_students
            total_students=len(records),
            present_count=len(present),
            absent_count=len(absent),
            records=records,
        )
        return data

    except Exception as e:
        logger.error(f"load_report_data error: {e}")
        return None


# ─────────────────────────────────────────────
#  Excel Report
# ─────────────────────────────────────────────
def export_excel(data: ReportData, output_path: str = None) -> Optional[str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        C_DARK_BG  = "07090F"; C_HEADER   = "0A1628"; C_CYAN    = "06C8E8"
        C_GREEN    = "10D98A"; C_RED      = "F04060"; C_ORANGE  = "F59E0B"
        C_TEXT     = "DCE8F8"; C_DIM      = "4A6080"; C_PRESENT = "0D3D24"
        C_ABSENT   = "3D0D14"; C_TITLE_BG = "050D1E"

        def border_all():
            s = Side(style="thin", color="1C2E4A")
            return Border(left=s, right=s, top=s, bottom=s)

        # ── Group by Class ──
        records_by_class = {}
        for r in (data.records or []):
            c_name = r.get("class_name", "Khác")
            if c_name not in records_by_class:
                records_by_class[c_name] = []
            records_by_class[c_name].append(r)
            
        if not records_by_class:
            records_by_class["Khác"] = []

        is_first_sheet = True
        for c_name, c_records in records_by_class.items():
            if is_first_sheet:
                ws1 = wb.active
                ws1.title = c_name[:31]
                is_first_sheet = False
            else:
                ws1 = wb.create_sheet(c_name[:31])
            
            ws1.sheet_view.showGridLines = False
            ws1.sheet_properties.tabColor = C_CYAN

            col_widths = [5, 12, 30, 10, 14, 14, 14, 12, 12, 14]
            for i, w in enumerate(col_widths, 1):
                ws1.column_dimensions[get_column_letter(i)].width = w

            ws1.merge_cells("B1:K1")
            c = ws1["B1"]
            c.value = "HỆ THỐNG ĐIỂM DANH KHUÔN MẶT"
            c.font = Font(name="Arial", size=18, bold=True, color=C_CYAN)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor=C_TITLE_BG)
            ws1.row_dimensions[1].height = 40

            ws1.merge_cells("B2:K2")
            c = ws1["B2"]
            c.value = data.title
            c.font = Font(name="Arial", size=14, bold=True, color=C_TEXT)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor=C_HEADER)
            ws1.row_dimensions[2].height = 30

            info_rows = [
                ("Lớp học:",   f"{c_name}"),
                ("Môn học:",   data.subject_name),
                ("Ngày:",      data.session_date),
                ("Bắt đầu:",   data.start_time),
                ("Kết thúc:",  data.end_time),
                ("Giáo viên:", data.teacher_name or "—"),
            ]
            for i, (label, value) in enumerate(info_rows, 4):
                ws1.row_dimensions[i].height = 22
                ws1.merge_cells(f"B{i}:C{i}")
                lbl = ws1[f"B{i}"]
                lbl.value = label
                lbl.font = Font(name="Arial", size=10, bold=True, color="8BA4C0")
                lbl.fill = PatternFill("solid", fgColor=C_HEADER)
                lbl.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                ws1.merge_cells(f"D{i}:G{i}")
                val = ws1[f"D{i}"]
                val.value = value
                val.font = Font(name="Arial", size=11, bold=True, color=C_TEXT)
                val.fill = PatternFill("solid", fgColor=C_HEADER)
                val.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            stat_row = 11
            c_total = len(c_records)
            c_present = len([r for r in c_records if r.get("status") == "PRESENT"])
            c_absent = len([r for r in c_records if r.get("status") == "ABSENT"])
            c_rate = (c_present / c_total * 100) if c_total > 0 else 0.0

            stats = [
                ("TỔNG HV",  c_total,              C_CYAN),
                ("CÓ MẶT",   c_present,               C_GREEN),
                ("VẮNG MẶT", c_absent,                C_RED),
                ("TỈ LỆ",    f"{c_rate:.1f}%",   C_ORANGE),
            ]
            for (label, value, color), col in zip(stats, ["B","D","F","H"]):
                ec = chr(ord(col)+1)
                if col == "H":
                    ws1.merge_cells(f"H{stat_row}:K{stat_row}")
                    ws1.merge_cells(f"H{stat_row+1}:K{stat_row+1}")
                else:
                    ws1.merge_cells(f"{col}{stat_row}:{ec}{stat_row}")
                    ws1.merge_cells(f"{col}{stat_row+1}:{ec}{stat_row+1}")
                
                lbl = ws1[f"{col}{stat_row}"]
                lbl.value = label
                lbl.font = Font(name="Arial", size=9, bold=True, color=color)
                lbl.alignment = Alignment(horizontal="center", vertical="center")
                lbl.fill = PatternFill("solid", fgColor=C_HEADER)
                ws1.row_dimensions[stat_row].height = 22
                val = ws1[f"{col}{stat_row+1}"]
                val.value = value
                val.font = Font(name="Arial", size=20, bold=True, color=color)
                val.alignment = Alignment(horizontal="center", vertical="center")
                val.fill = PatternFill("solid", fgColor=f"{color}18")
                ws1.row_dimensions[stat_row+1].height = 38

            tbl_start = stat_row + 3
            headers = ["STT","Mã HV","Họ và Tên","Giới tính",
                       "Trạng thái","Giờ điểm danh","Độ chính xác", "Tòa nhà", "Phòng", "Camera"]
            for j, h in enumerate(headers, 2):
                c = ws1.cell(row=tbl_start, column=j)
                c.value = h
                c.font = Font(name="Arial", size=10, bold=True, color=C_TEXT)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill("solid", fgColor=C_HEADER)
                c.border = border_all()
            ws1.row_dimensions[tbl_start].height = 28

            for idx, record in enumerate(c_records, 1):
                row = tbl_start + idx
                ws1.row_dimensions[row].height = 22
                is_p = record.get("status") == "PRESENT"
                bg   = C_PRESENT if is_p else C_ABSENT
                txt  = C_GREEN   if is_p else C_RED
                vals = [
                    idx,
                    record.get("student_code", ""),
                    record.get("full_name", ""),
                    record.get("gender", ""),
                    "✓ Có mặt" if is_p else "✗ Vắng",
                    record.get("check_in_time", ""),
                    f"{record.get('recognition_score',0)*100:.1f}%" if is_p else "—",
                    record.get("building", ""),
                    record.get("room", ""),
                    f"Cam {record.get('camera_id','')}" if is_p else "—",
                ]
                for j, v in enumerate(vals, 2):
                    c = ws1.cell(row=row, column=j)
                    c.value = v
                    c.font = Font(name="Arial", size=10,
                                  bold=(j==4),
                                  color=txt if j==6 else C_TEXT)
                    c.alignment = Alignment(
                        horizontal="left" if j==4 else "center",
                        vertical="center", indent=1 if j==4 else 0)
                    c.fill = PatternFill("solid", fgColor=bg)
                    c.border = border_all()

            footer_row = tbl_start + len(c_records) + 2
            ws1.merge_cells(f"B{footer_row}:K{footer_row}")
            f = ws1[f"B{footer_row}"]
            f.value = (
                f"Báo cáo tạo lúc: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} "
                f"| Session ID: {data.session_id}"
            )
            f.font = Font(name="Arial", size=9, italic=True, color=C_DIM)
            f.alignment = Alignment(horizontal="center")

        if not output_path:
            fname = (
                f"BaoCao_{data.class_code}_{data.subject_name[:20]}_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            ).replace("/","-").replace("\\","-").replace(" ","_")
            output_path = str(report_config.output_dir / fname)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.success(f"Excel saved: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"export_excel error: {e}")
        return None


# ─────────────────────────────────────────────
#  PDF Report
# ─────────────────────────────────────────────
def export_pdf(data: ReportData, output_path: str = None) -> Optional[str]:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable,
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        if not output_path:
            fname = (
                f"BaoCao_{data.class_code}_{data.subject_name[:20]}_"
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            ).replace("/","-").replace("\\","-").replace(" ","_")
            output_path = str(report_config.output_dir / fname)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        CLR_PANEL   = colors.HexColor("#0D1320"); CLR_CARD  = colors.HexColor("#111B2E")
        CLR_CYAN    = colors.HexColor("#06C8E8"); CLR_GREEN = colors.HexColor("#10D98A")
        CLR_RED     = colors.HexColor("#F04060"); CLR_TEXT  = colors.HexColor("#DCE8F8")
        CLR_DIM     = colors.HexColor("#8BA4C0"); CLR_DARK  = colors.HexColor("#4A6080")

        doc = SimpleDocTemplate(
            output_path, pagesize=A4,
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=1.5*cm,  bottomMargin=1.5*cm,
        )

        def sty(name, **kw):
            return ParagraphStyle(name, **kw)

        S_TITLE   = sty("t", fontSize=18, fontName="Helvetica-Bold",
                        textColor=CLR_CYAN, alignment=TA_CENTER, spaceAfter=4)
        S_SUB     = sty("s", fontSize=13, fontName="Helvetica-Bold",
                        textColor=CLR_TEXT, alignment=TA_CENTER, spaceAfter=8)
        S_SECTION = sty("sec", fontSize=11, fontName="Helvetica-Bold",
                        textColor=CLR_CYAN, alignment=TA_LEFT, spaceBefore=8)
        S_FOOTER  = sty("f", fontSize=8, fontName="Helvetica",
                        textColor=CLR_DARK, alignment=TA_CENTER)
        S_CELL    = sty("c",  fontSize=9, fontName="Helvetica",
                        textColor=CLR_TEXT, alignment=TA_CENTER)
        S_CELL_L  = sty("cl", fontSize=9, fontName="Helvetica",
                        textColor=CLR_TEXT, alignment=TA_LEFT)

        W = A4[0] - 3*cm
        story = []

        story.append(Paragraph("HỆ THỐNG ĐIỂM DANH KHUÔN MẶT", S_TITLE))
        story.append(Paragraph(data.title, S_SUB))
        story.append(HRFlowable(width="100%", thickness=1, color=CLR_CYAN, spaceAfter=8))

        info = [
            ["Lớp học:", f"{data.class_code} — {data.class_name}", "Ngày:",    data.session_date],
            ["Môn học:", data.subject_name,                         "Giờ:",     f"{data.start_time} — {data.end_time}"],
            ["Giáo viên:", data.teacher_name or "—",               "Session:", str(data.session_id)],
        ]
        info_tbl = Table(info, colWidths=[2.2*cm, 7*cm, 2*cm, 4*cm])
        info_tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,-1), CLR_PANEL),
            ("TEXTCOLOR",    (0,0), (0,-1),  CLR_DIM),
            ("TEXTCOLOR",    (2,0), (2,-1),  CLR_DIM),
            ("TEXTCOLOR",    (1,0), (1,-1),  CLR_TEXT),
            ("TEXTCOLOR",    (3,0), (3,-1),  CLR_TEXT),
            ("FONTNAME",     (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 9),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("LEFTPADDING",  (0,0), (-1,-1), 6),
            ("GRID",         (0,0), (-1,-1), 0.3, CLR_DARK),
        ]))
        story.append(info_tbl)
        story.append(Spacer(1, 12))

        rc = "#10D98A" if data.attendance_rate >= 80 else \
             "#F59E0B" if data.attendance_rate >= 60 else "#F04060"
        stats_data = [[
            Paragraph(f"<font color='#8BA4C0' size='8'>TỔNG HỌC VIÊN</font><br/>"
                      f"<font color='#06C8E8' size='22'><b>{data.total_students}</b></font>", S_CELL),
            Paragraph(f"<font color='#8BA4C0' size='8'>CÓ MẶT</font><br/>"
                      f"<font color='#10D98A' size='22'><b>{data.present_count}</b></font>", S_CELL),
            Paragraph(f"<font color='#8BA4C0' size='8'>VẮNG MẶT</font><br/>"
                      f"<font color='#F04060' size='22'><b>{data.absent_count}</b></font>", S_CELL),
            Paragraph(f"<font color='#8BA4C0' size='8'>TỈ LỆ</font><br/>"
                      f"<font color='{rc}' size='22'><b>{data.attendance_rate:.1f}%</b></font>", S_CELL),
        ]]
        st = Table(stats_data, colWidths=[W/4]*4)
        st.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,-1), CLR_CARD),
            ("TOPPADDING",   (0,0), (-1,-1), 10),
            ("BOTTOMPADDING",(0,0), (-1,-1), 10),
            ("LINEABOVE",    (0,0), (-1,0),  1, CLR_CYAN),
            ("LINEBELOW",    (0,-1),(-1,-1), 1, CLR_CYAN),
            ("INNERGRID",    (0,0), (-1,-1), 0.5, CLR_DARK),
            ("ALIGN",        (0,0), (-1,-1), "CENTER"),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(st)
        story.append(Spacer(1, 14))

        story.append(Paragraph("BẢNG ĐIỂM DANH CHI TIẾT", S_SECTION))
        story.append(Spacer(1, 6))

        hdrs = [Paragraph(f"<b>{h}</b>", S_CELL)
                for h in ["STT","Mã HV","Họ và Tên","Trạng thái","Giờ điểm danh","Độ chính xác"]]
        tbl_data = [hdrs]
        for idx, r in enumerate(data.records or [], 1):
            is_p = r.get("status") == "PRESENT"
            st_txt = f'<font color="#10D98A">✓ Có mặt</font>' if is_p \
                     else f'<font color="#F04060">✗ Vắng</font>'
            tbl_data.append([
                Paragraph(str(idx), S_CELL),
                Paragraph(r.get("student_code",""), S_CELL),
                Paragraph(r.get("full_name",""), S_CELL_L),
                Paragraph(st_txt, S_CELL),
                Paragraph(str(r.get("check_in_time","—")), S_CELL),
                Paragraph(f"{r.get('recognition_score',0)*100:.1f}%" if is_p else "—", S_CELL),
            ])

        dt = Table(tbl_data, colWidths=[1*cm,2.2*cm,6*cm,2.8*cm,3*cm,2.5*cm], repeatRows=1)
        dt.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  CLR_PANEL),
            ("LINEBELOW",     (0,0), (-1,0),  1.5, CLR_CYAN),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [CLR_CARD, CLR_PANEL]),
            ("GRID",          (0,0), (-1,-1), 0.3, CLR_DARK),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 4),
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN",         (2,1), (2,-1),  "LEFT"),
        ]))
        story.append(dt)
        story.append(Spacer(1, 20))

        story.append(HRFlowable(width="100%", thickness=0.5, color=CLR_DARK))
        story.append(Spacer(1, 4))
        story.append(Paragraph(
            f"Báo cáo tạo lúc {datetime.now().strftime('%H:%M:%S %d/%m/%Y')} "
            f"· Session ID: {data.session_id}",
            S_FOOTER
        ))

        doc.build(story)
        logger.success(f"PDF saved: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"export_pdf error: {e}")
        return None


# ─────────────────────────────────────────────
def generate_report(session_id: int, fmt: str = "both") -> dict:
    data = load_report_data(session_id)
    if not data:
        return {"success": False, "error": f"Không tìm thấy session {session_id}"}
    result = {"success": True, "excel": None, "pdf": None}
    if fmt in ("excel", "both"):
        result["excel"] = export_excel(data)
    if fmt in ("pdf", "both"):
        result["pdf"]   = export_pdf(data)
    if not any([result["excel"], result["pdf"]]):
        result["success"] = False
        result["error"]   = "Xuất file thất bại"
    return result